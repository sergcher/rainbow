import os
from datetime import datetime

from django.db.models import Sum
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from main.models import Apartment, ApartmentDetail, ApartmentFee, Settings, ApartmentCounter


def export_client_bank():
    filename = f'44069_{datetime.today().strftime("%d%m%Y")}_1.txt'
    total = 0
    apartments = Apartment.objects.all()
    settings = Settings.objects.get(id=1)

    date_str = settings.month_to_date
    date_obj = datetime.strptime(date_str, '%d.%m.%Y')
    date_str = datetime.strftime(date_obj, '%m%Y')

    with open(filename, 'w') as f:
        for apartment in apartments:
            apartment_detail = ApartmentDetail.objects.get(serialNumber=apartment.serialNumber)
            apartment_fee = ApartmentFee.objects.get(serialNumber=apartment.serialNumber)
            fee = round(apartment_fee.total, 2)
            total += fee
            f.write(
                f'{apartment_detail.personalAccount}|{apartment.owner.strip()}'
                f'|Междуреченск, Шахтеров проспект, д. 55,{apartment.serialNumber}|11|'
                f'Квартплата|{date_str}||{int(fee * 100)}' + '\n')
        total = round(total, 2)
        f.write('=|106|' + str(int(total * 100)) + '\n')

    with open(filename, 'rb') as f:
        response = HttpResponse(f.read(), content_type='text/plain')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

    os.remove(filename)
    return response


def generate_excel_file(apartments):
    wb = Workbook()
    ws = wb.active
    headers = [
         '№', 'ФИО', 'Кол. зарег.', 'Кол. прожив.', 'Общ. площ.', 'Содерж. помещения', 'Эл/эн ОДН', 'ОДН Холодная вода',
         'ОДН Сточные воды', 'ОДН Горячая вода', 'Лифт', 'Содерж. помещ. итого', 'Обращение с ТКО', 'Электр.',
         'Отопление Гкал', 'Отопление, руб.', 'Гор. вода', 'Хол. вода', 'Водоотв.', 'Итого коммунальные услуги',
         'Начислено', 'Перерасчет', 'Сальдо Начало', 'Сальдо Конец', 'Оплачено', 'Пеня', 'Итого к оплате',
         'Горячая вода показания предыдущие', 'Горячая вода показания текущие', 'Горячая вода потребление',
         'Холодная вода показания предыдущие', 'Холодная вода показания текущие', 'Холодная вода потребление',
         'Электричество показания предыдущие', 'Электричество показания текущие', 'Электричество вода потребление',
         'Отведение сточных вод'
    ]

    ws.append(headers)

    ws.row_dimensions[1].height = 30

    column_widths = [10, 30, 12, 12, 12, 18, 12, 18, 18, 18, 12, 20, 12, 12, 12, 12, 12, 12, 20, 12,
                     12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12]
    for i, width in enumerate(column_widths):
        col_letter = get_column_letter(i + 1)
        ws.column_dimensions[col_letter].width = width

    for i, apartment in enumerate(apartments):
        apartment_detail = ApartmentDetail.objects.get(serialNumber=apartment.serialNumber)
        apartment_fee = ApartmentFee.objects.get(serialNumber=apartment.serialNumber)
        apartment_counter = ApartmentCounter.objects.get(serialNumber=apartment.serialNumber)
        data = [
            apartment.serialNumber,
            apartment.owner,
            apartment_detail.registredQt,
            apartment_detail.livedQt,
            apartment_detail.totalArea,
            apartment_fee.maintenance,
            apartment_fee.electricity_odn,
            apartment_fee.cold_water_odn,
            apartment_fee.sewage_odn,
            apartment_fee.hot_water_odn,
            apartment_fee.lift,
            apartment_fee.maintenance_full,
            apartment_fee.solid_waste,
            apartment_fee.electricity,
            apartment_fee.heating,
            apartment_fee.heating_rub,
            apartment_fee.hot_water,
            apartment_fee.cold_water,
            apartment_fee.sewage,
            apartment_fee.maintenance_total,
            apartment_fee.accrued_expenses,
            apartment_fee.recalculation,
            apartment_fee.balance_start,
            apartment_fee.balance_end,
            apartment_fee.paid,
            apartment_fee.fine,
            apartment_fee.total,
            apartment_counter.hot_water_previous,
            apartment_counter.hot_water_current,
            apartment_counter.hot_water_value,
            apartment_counter.cold_water_previous,
            apartment_counter.cold_water_current,
            apartment_counter.cold_water_value,
            apartment_counter.electricity_previous,
            apartment_counter.electricity_current,
            apartment_counter.electricity_value,
            apartment_counter.wastewater_value
        ]
        ws.append(data)

    for column in range(1, len(headers) + 1):
        col_letter = get_column_letter(column)
        cell = ws["{}{}".format(col_letter, i + 2)]
        cell.alignment = Alignment(wrap_text=True)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=apartment_fees.xlsx'
    wb.save(response)

    return response


def export_excel_apartment_total_file():
    wb = Workbook()
    ws = wb.active

    # Кол-во прописанных
    value = ApartmentDetail.objects.aggregate(Sum('registredQt'))['registredQt__sum']
    data = ['Кол-во прописанных', value]
    ws.append(data)

    # Кол-во проживающих
    value = ApartmentDetail.objects.aggregate(Sum('livedQt'))['livedQt__sum']
    data = ['Кол-во проживающих', value]
    ws.append(data)

    # Общая площадь
    value = ApartmentDetail.objects.aggregate(Sum('totalArea'))['totalArea__sum']
    data = ['Общая площадь', value]
    ws.append(data)

    # Отапливаемая площадь
    value = ApartmentDetail.objects.aggregate(Sum('totalArea'))['totalArea__sum']
    data = ['Отапливаемая площадь', value]
    ws.append(data)

    # Электроэнергия
    value = ApartmentFee.objects.aggregate(Sum('electricity'))['electricity__sum']
    data = ['Электроэнергия', value]
    ws.append(data)

    # Отопление
    value = ApartmentFee.objects.aggregate(Sum('heating_rub'))['heating_rub__sum']
    data = ['Отопление', value]
    ws.append(data)

    # Горячая вода
    value = ApartmentFee.objects.aggregate(Sum('hot_water'))['hot_water__sum']
    data = ['Горячая вода', value]
    ws.append(data)

    # Холодная вода
    value = ApartmentFee.objects.aggregate(Sum('cold_water'))['cold_water__sum']
    data = ['Холодная вода', value]
    ws.append(data)

    # Водоотведение
    value = ApartmentFee.objects.aggregate(Sum('sewage'))['sewage__sum']
    data = ['Водоотведение', value]
    ws.append(data)

    # Обращение с ТКО
    value = ApartmentFee.objects.aggregate(Sum('solid_waste'))['solid_waste__sum']
    data = ['Обращение с ТКО', value]
    ws.append(data)

    # Содержание помещения
    value = ApartmentFee.objects.aggregate(Sum('maintenance'))['maintenance__sum']
    data = ['Содержание помещения', value]
    ws.append(data)

    # ОДН Эл/энергия
    value = ApartmentFee.objects.aggregate(Sum('electricity_odn'))['electricity_odn__sum']
    data = ['ОДН Эл/энергия', value]
    ws.append(data)

    # ОДН сточные воды
    value = ApartmentFee.objects.aggregate(Sum('sewage_odn'))['sewage_odn__sum']
    data = ['ОДН сточные воды', value]
    ws.append(data)

    # ОДН холодная вода
    value = ApartmentFee.objects.aggregate(Sum('cold_water_odn'))['cold_water_odn__sum']
    data = ['ОДН холодная вода', value]
    ws.append(data)

    # ОДН горячая вода
    value = ApartmentFee.objects.aggregate(Sum('hot_water_odn'))['hot_water_odn__sum']
    data = ['ОДН горячая вода', value]
    ws.append(data)

    # Лифт
    value = ApartmentFee.objects.aggregate(Sum('lift'))['lift__sum']
    data = ['Лифт', value]
    ws.append(data)

    # Сальдо конец
    value = ApartmentFee.objects.aggregate(Sum('balance_end'))['balance_end__sum']
    data = ['Сальдо конец', value]
    ws.append(data)

    # Оплачено
    value = ApartmentFee.objects.aggregate(Sum('paid'))['paid__sum']
    data = ['Оплачено', value]
    ws.append(data)

    # Сальдо начало
    value = ApartmentFee.objects.aggregate(Sum('balance_start'))['balance_start__sum']
    data = ['Сальдо начало', value]
    ws.append(data)

    # Пеня
    value = ApartmentFee.objects.aggregate(Sum('fine'))['fine__sum']
    data = ['Пеня', value]
    ws.append(data)

    # Перерасчет/недопоставка услуг
    value = ApartmentFee.objects.aggregate(Sum('recalculation'))['recalculation__sum']
    data = ['Перерасчет/ недопоставка услуг', value]
    ws.append(data)

    # Начислено
    value = ApartmentFee.objects.aggregate(Sum('accrued_expenses'))['accrued_expenses__sum']
    data = ['Начислено', value]
    ws.append(data)

    # Итого
    value = ApartmentFee.objects.aggregate(Sum('total'))['total__sum']
    data = ['Итого', value]
    ws.append(data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=apartment_fees.xlsx'
    wb.save(response)

    return response
